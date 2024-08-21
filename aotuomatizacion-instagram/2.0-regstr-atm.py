import datetime
from lib2to3.pgen2 import driver
import os
from pyclbr import Class
import random
import re
from ssl import Options
import threading
import time
from weakref import ProxyType
from httpcore import TimeoutException
from httpx import Proxy
from openpyxl import Workbook, load_workbook
import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
import faker
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.proxy import Proxy, ProxyType

fake = faker.Faker()

class NotepadCreator:
    def __init__(self, filename):
        self.filename = filename
        # Crear o sobrescribir el archivo al iniciar
        with open(self.filename, 'w') as file:
            file.write("")  # Vacía el archivo o lo crea si no existe

    def add_data(self, correo, nombre, usuario):
        with open(self.filename, 'a') as file:
            file.write(f"{correo}:{nombre}:{usuario}\n")

    def save(self):
        print(f"Archivo '{self.filename}' guardado exitosamente.")



class CodigoVerificacionGmail:
    def __init__(self, email):
        self.email = email
        self.username = email.split('@')[0]
        self.domain = "1secmail.com"
    
    def check_inbox(self):
        url = f"https://www.1secmail.com/api/v1/?action=getMessages&login={self.username}&domain={self.domain}"
        response = requests.get(url)
        messages = response.json()
        
        if messages:
            for message in messages:
                # Filtrar mensajes de verificación de Instagram
                if 'instagram' in message['subject'].lower():
                    print(f"Correo de verificación encontrado con ID: {message['id']}")
                    return message['id']
        return None
    
    def get_message(self, message_id):
        url = f"https://www.1secmail.com/api/v1/?action=readMessage&login={self.username}&domain={self.domain}&id={message_id}"
        response = requests.get(url)
        message = response.json()
        
        # Verifica las posibles claves donde podría estar el contenido del mensaje
        if 'text' in message:
            return message['text']
        elif 'body' in message:
            return message['body']
        else:
            raise KeyError("El contenido del mensaje no se encuentra en las claves esperadas.")
    
    def extract_verification_code(self, message_text):
        # Utilizar expresión regular para encontrar un código de verificación en el correo
        match = re.search(r'\b\d{6}\b', message_text)  # Ajusta el patrón según el formato del código
        if match:
            return match.group(0)
        return None
    
    def esperar_verificacion(self):
        print("Esperando correo de verificación...")
        while True:
            message_id = self.check_inbox()
            if message_id:
                message_content = self.get_message(message_id)
                verification_code = self.extract_verification_code(message_content)
                if verification_code:
                    print(f"Código de verificación recibido: {verification_code}")
                    return verification_code
                else:
                    print("Código de verificación no encontrado en el mensaje.")
                break
            time.sleep(5)  # Esperar 30 segundos antes de volver a comprobar
    


# Interface para la autenticación
class AutenticacionInterface:
    def gnercin_nmbr_apel_ramdom(self):
        pass


    def ingresar_correo(self, correo):
        pass


    def ingresar_nombre_completo(self, nombreCom):
        pass


    def ingresar_nombre_intagram_prederter(self, nombreInstagram):
        pass


    def ingresar_contraseña(self, contrase):
        pass

    
    def hacer_clic_en_boton(self):
        pass

    
    def verificar_cambio_de_pagina(self):
        pass


    def genereador_fecha_nacimiento(self):
        pass


    def ingresar_mes(self):        
        pass


    def ingresar_dia(self):       
        pass


    def ingresar_año(self):         
        pass
    
    
    def hacer_clic_en_boton_fecha(self):        
        pass


    def mantener_sesion(self):
        pass


    def vericacion_codigo_instagram(self):
        pass


    def esperar_inicio_seccion(self):
        pass


    def iniciar_session_correo(self):
        pass


    def iniciar_session_contraseña(self):
        pass


    def boton_iniciar_session(self):
        pass

    def boton_redireecionar_inicio(self):
        pass

    def guardar_datos_excel(self):
        pass
    

    

# Clase que implementa la autenticación
class Autenticacion(AutenticacionInterface):
    def __init__(self, driver):
        self.driver = driver

    # Generar nombre, apellido y correo de manera aleatoria
    def gnercin_nmbr_apel_ramdom(self):
        nombre = fake.first_name()
        apellido = fake.last_name()
        nombre_apel = f"{nombre.lower()}{" "}{apellido.lower()}"
        correo = f"{nombre.lower()}{apellido.lower()}@1secmail.com" 
        # Devolver las variables separadas
        return   nombre_apel,correo
    
    
    def ingresar_correo(self, correo):
        username_input = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.NAME, "emailOrPhone"))
        )
        for char in correo:
            username_input.send_keys(char)
            time.sleep(0.1)


    def ingresar_nombre_completo(self, nombreCom):
        password_input = self.driver.find_element(By.NAME, "fullName")
        for char in nombreCom:
    
            password_input.send_keys(char)
            time.sleep(0.3)


    def ingresar_nombre_intagram_prederter(self):
        try:
            # Esperar hasta que el botón "Actualizar sugerencia" sea visible y esté habilitado
            button = WebDriverWait(self.driver, 5).until(
                EC.element_to_be_clickable((By.XPATH, "//button[span[text()='Actualizar sugerencia']]"))               
            )

            # Hacer clic en el botón
            button.click()

        except TimeoutException:
            print("El botón 'Actualizar sugerencia' no se encontró en el tiempo especificado.")
        except Exception as e:
            print(f"Error al intentar hacer clic en el botón 'Actualizar sugerencia': {e}")


    def ingresar_contraseña(self,contrase):
        password_input = self.driver.find_element(By.NAME, "password")
        for char in contrase:
            password_input.send_keys(char)
            time.sleep(0.1)

    
    def hacer_clic_en_boton(self):
        try:
            # Esperar hasta que el botón "Registrarte" sea visible y esté habilitado
            button = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, '//button[contains(@class, "_acan _acap _acas _aj1- _ap30") and @type="submit" and text()="Registrarte"]'))
            )

            # Hacer clic en el botón
            button.click()

        except TimeoutException:
            print("El botón 'Registrarte' no se encontró en el tiempo especificado.")
        except Exception as e:
            print(f"Error al intentar hacer clic en el botón 'Registrarte': {e}")
    
    def verificar_cambio_de_pagina(self):
        try:
            # Espera hasta que el nuevo elemento con el span específico esté presente en la página
            WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'span.x1lliihq.x1plvlek.xryxfnj.x1n2onr6.x193iq5w.xeuugli.x1fj9vlw.x13faqbe.x1vvkbs.x1s928wv.xhkezso.x1gmr53x.x1cpjm7i.x1fgarty.x1943h6x.x1i0vuye.xvs91rp.x1s688f.x5n08af.x2b8uid.x1tu3fi.x3x7a5m.x10wh9bi.x1wdrske.x8viiok.x18hxmgj[dir="auto"]'))
            )
            print("SE HA INGRESADO CORRECTAMENTE, AHORA SE SEGUIRÁ CON EL REGISTRO DE FECHA")
        
        except Exception as e:
            print("No se ha producido cambio de página.", e)

    def genereador_fecha_nacimiento(self):
        # Generar un número aleatorio para el día entre 1 y 31
        dia = random.randint(1, 31)
        
        # Generar un número aleatorio para el mes entre 1 y 12
        mes = random.randint(1, 12)
        
        # Generar un número aleatorio para el año entre 1994 y 2004
        año = random.randint(1994, 2004)
    
        return  dia, mes, año
    
    def seleccionar_fecha_nacimiento(self, dia, mes, año):
        Select(self.driver.find_element(By.CSS_SELECTOR, 'select[title="Mes:"]')).select_by_value(str(mes))
        Select(self.driver.find_element(By.CSS_SELECTOR, 'select[title="Día:"]')).select_by_value(str(dia))
        Select(self.driver.find_element(By.CSS_SELECTOR, 'select[title="Año:"]')).select_by_value(str(año))

    
    def hacer_clic_en_boton_fecha(self):
        try:
            # Esperar hasta que el botón "Siguiente" sea visible
            button = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, '//button[contains(@class, "_acan _acap _acaq _acas _aj1- _ap30") and text()="Siguiente"]'))
            )

            # Esperar hasta que el botón esté habilitado
            WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, '//button[contains(@class, "_acan _acap _acaq _acas _aj1- _ap30") and text()="Siguiente" and not(@disabled)]'))
            )

            # Hacer clic en el botón
            button.click()

        except TimeoutException:
            print("El botón 'Siguiente' no se encontró en el tiempo especificado o no se habilitó.")
        except Exception as e:
            print(f"Error al intentar hacer clic en el botón 'Siguiente': {e}")

    def mantener_sesion(self, duracion_segundos):
        try:
            print(f"Manteniendo la sesión activa durante {duracion_segundos} segundos...")
            time.sleep(duracion_segundos)
        except KeyboardInterrupt:
            print("Sesión terminada manualmente.")
        finally:
            self.cerrar_navegador()
            print("Navegador cerrado.")

    def vericacion_codigo_instagram(self,codigo):
        password_input = self.driver.find_element(By.NAME, "email_confirmation_code")
        for char in codigo:
            password_input.send_keys(char)
            time.sleep(0.1)
    

    def presionar_botton_verificacion(self):
        try:
            # Esperar hasta que el botón "Siguiente" sea visible
            button = WebDriverWait(self.driver, 10).until(
                   
                EC.element_to_be_clickable((By.XPATH, "//div[@role='button' and text()='Siguiente']"))
            )
            # Hacer clic en el botón
            button.click()

        except TimeoutException:
            print("El botón 'Siguiente' no se encontró en el tiempo especificado o no se habilitó.")
        except Exception as e:
            print(f"Error al intentar hacer clic en el botón 'Siguiente': {e}")
    
    def esperar_inicio_seccion(self):
         # Esperar hasta que el inicio de sesión sea exitoso (verificar con la presencia de un elemento del dashboard)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//div[text()='Inicio']")))
        print("Inicio de sesión exitoso.")
    
    
    def iniciar_session_correo(self,correo):
        username_input = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.NAME, "username"))
        )
        for char in correo:
            username_input.send_keys(char)
            time.sleep(0.1)


    def iniciar_session_contraseña(self, contaseña):
        username_input = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.NAME, "password"))
        )
        for char in contaseña:
            username_input.send_keys(char)
            time.sleep(0.1)
        


    def boton_iniciar_session(self):
        try:
            # Esperar hasta que el botón "Siguiente" sea visible
            button = WebDriverWait(self.driver, 10).until(              
            EC.element_to_be_clickable((By.XPATH, "//button[div[contains(text(), 'Entrar')]]"))
            )
            # Hacer clic en el botón
            button.click()

        except TimeoutException:
            print("El botón 'Siguiente' no se encontró en el tiempo especificado o no se habilitó.")
        except Exception as e:
            print(f"Error al intentar hacer clic en el botón 'Siguiente': {e}")

    
    def boton_redireecionar_inicio(self):
        try:
            # Esperar hasta que el botón "Siguiente" sea visible
            button = WebDriverWait(self.driver, 10).until(              
            boton = driver.find_element(By.CSS_SELECTOR, 'div.x9f619.xxk0z11.x11xpdln.xvy4d1p.x1o5bo1o.x1eub6wo.x19c4wfv svg[aria-label="Inicio"]')
            )
            # Hacer clic en el botón
            button.click()

        except TimeoutException:
            print("El botón 'Siguiente' no se encontró en el tiempo especificado o no se habilitó.")
        except Exception as e:
            print(f"Error al intentar hacer clic en el botón 'Siguiente': {e}")


    def guardar_datos_excel(self, correo, nombre_usuario):
        archivo_excel = "registros_instagram.xlsx"

        # Si el archivo no existe, se crea uno nuevo
        if not os.path.exists(archivo_excel):
            wb = Workbook()
            ws = wb.active
            ws.append(["Correo", "Nombre Usuario", "Fecha"])  # Agregar cabeceras
            wb.save(archivo_excel)

        # Cargar el archivo existente y añadir una nueva fila
        try:
            wb = load_workbook(archivo_excel)
        except Exception as e:
            print(f"Error al cargar el archivo {archivo_excel}: {e}")
            return

        ws = wb.active

        # Obtener la fecha actual
        fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Añadir los datos a la siguiente fila disponible
        try:
            ws.append([correo, nombre_usuario, fecha_actual])
        except Exception as e:
            print(f"Error al agregar datos a la fila: {e}")
            return

        # Guardar los cambios en el archivo
        try:
            wb.save(archivo_excel)
        except Exception as e:
            print(f"Error al guardar el archivo {archivo_excel}: {e}")
            return

        print(f"Datos guardados en {archivo_excel}: {correo}, {nombre_usuario}, {fecha_actual}")


# Clase que configura el navegador
class Navegador:
    def __init__(self, proxy=None):
        chrome_options = Options()

        # # Activar modo incógnito para mayor anonimato
        # chrome_options.add_argument("--incognito")

        # Deshabilitar el rastreo de automatización del navegador
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_argument("--disable-infobars")

        # Opciones adicionales
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        
        # Configurar proxy si es proporcionado
        if proxy:
            chrome_proxy = Proxy()
            chrome_proxy.proxy_type = ProxyType.MANUAL
            chrome_proxy.http_proxy = proxy
            chrome_proxy.ssl_proxy = proxy
            chrome_options.proxy = chrome_proxy
            chrome_options.add_argument(f'--proxy-server={proxy}')

        # Inicializar el driver de Chrome con las opciones configuradas
        self.driver = webdriver.Chrome(options=chrome_options)

    def abrir_pagina(self, url):
        self.driver.get(url)

    def obtener_cookies(self):
        # Obtener las cookies de sesión
        cookies = self.driver.get_cookies()
        return cookies
    

    def cerrar_navegador(self):
        self.driver.quit()


class ProxyTester:
    def __init__(self, ruta_archivo_proxies, url_prueba):
        self.ruta_archivo_proxies = ruta_archivo_proxies
        self.url_prueba = url_prueba
        self.proxies = self.cargar_proxies_desde_archivo()
        self.proxies_validos = []

    def cargar_proxies_desde_archivo(self):
        proxies = []
        try:
            with open(self.ruta_archivo_proxies, 'r') as archivo:
                for linea in archivo:
                    linea = linea.strip()
                    if linea:
                        proxy = {"http": f"http://{linea}", "https": f"http://{linea}"}
                        proxies.append(proxy)
        except FileNotFoundError:
            print(f"Error: No se encontró el archivo {self.ruta_archivo_proxies}")
        return proxies

    def probar_proxy(self, proxy):
        try:
            response = requests.get(self.url_prueba, proxies=proxy, timeout=5)
            ip = response.json()['origin']
            proxy_ip = proxy['http'].split(':')[1][2:]
            print(f"Proxy válido: {proxy_ip} - IP: {ip}")
            self.proxies_validos.append(proxy)
        except requests.exceptions.RequestException:
            pass

    def probar_proxies(self):
        threads = []
        for proxy in self.proxies:
            thread = threading.Thread(target=self.probar_proxy, args=(proxy,))
            thread.start()
            threads.append(thread)
        
        for thread in threads:
            thread.join()

    def obtener_siguiente_proxy(self):
        if self.proxies_validos:
            return self.proxies_validos.pop(0)  # Sacar el primer proxy válido de la lista
        else:
            self.probar_proxies()  # Volver a probar si ya se han usado todos
            return self.obtener_siguiente_proxy() if self.proxies_validos else None



# Clase que coordina la autenticación
class CoordinadorAutenticacion:
    def __init__(self, navegador, autenticacion,proxy_tester,arhivoDatos):
        self.navegador = navegador
        self.autenticacion = autenticacion
        self.proxy_tester = proxy_tester
        self.arhivoDatos = arhivoDatos

    def autenticar(self, contraseña):
        while True:
            # Obtener el siguiente proxy válido
            proxy = self.proxy_tester.obtener_siguiente_proxy()
            if proxy is None:
                print("No hay proxies válidos disponibles.")
                break

            # Usar el proxy para la autenticación
            navegador = Navegador(proxy=proxy['http'])

            try:
                # Abrir la página de registro de Instagram
                self.navegador.abrir_pagina("https://www.instagram.com/accounts/emailsignup/")
                
                # Obtener los datos generados (nombre, apellido, correo)
                nombreCom, correo = self.autenticacion.gnercin_nmbr_apel_ramdom()
                
                # Ingreso del imputs automaticamente
                self.autenticacion.ingresar_correo(correo)
                self.autenticacion.ingresar_nombre_completo(nombreCom)

                # Actualizar sugerencia de nombre de usuario
                self.autenticacion.ingresar_nombre_intagram_prederter()

                # Ingreso de la contraseña
                self.autenticacion.ingresar_contraseña(contraseña)

                # Hacer clic en el botón de registro
                self.autenticacion.hacer_clic_en_boton()

                # Verificar que se ha cambiado de página para confirmar el registro
                self.autenticacion.verificar_cambio_de_pagina()

                # Generar fecha de nacimiento aleatoria
                dia, mes, año = self.autenticacion.genereador_fecha_nacimiento()

                # Ingresar fecha de nacimiento
                self.autenticacion.seleccionar_fecha_nacimiento(dia, mes, año)

                # Hacer clic en el botón de fecha
                self.autenticacion.hacer_clic_en_boton_fecha()

                # Esperar y obtener el código de verificación
                codigo_verificacion = CodigoVerificacionGmail(correo)
                codigo = codigo_verificacion.esperar_verificacion()

                print(f"El código de verificación de Instagram es: {codigo}")

                # Ingresar el código de verificación
                self.autenticacion.vericacion_codigo_instagram(codigo)

                # Hacer clic en el botón de verificación
                self.autenticacion.presionar_botton_verificacion()

                # Guardar los datos en el archivo
                arhivoDatos = NotepadCreator('usuarios.txt')
                self.arhivoDatos.add_data(correo, nombreCom, nombreCom)
                self.arhivoDatos.save()

                # Esperar 5 segundos
                time.sleep(5)

                # # Abrir inicio de sesión en Instagram
                # self.navegador.abrir_pagina("https://www.instagram.com/")

                # # Ingresar correo y contraseña
                # self.autenticacion.iniciar_session_correo(correo)
                # self.autenticacion.iniciar_session_contraseña(contraseña)

                # # Hacer clic en el botón de inicio de sesión
                # self.autenticacion.boton_iniciar_session()

                # # Espera de 8 segundos
                # time.sleep(8)

                # # Redireccionar a inicio
                # self.autenticacion.boton_redireecionar_inicio()

                # # Obtener la URL actual
                # current_url = self.navegador.driver.current_url
                # print("URL actual:", current_url)

                # Esperar 6 segundos
                time.sleep(6)

                # cerrar navegador
                self.navegador.cerrar_navegador()
                continue
            except Exception as e:
                print("Fallo con proxy {proxy['http']}")
                continue  # Intentar con el siguiente proxy en caso de fallo

# Uso
ruta_archivo_proxies = 'proxyes/proxies.txt'
url_prueba = "http://httpbin.org/ip"
proxy_tester = ProxyTester(ruta_archivo_proxies, url_prueba)

navegador = Navegador()
autenticacion = Autenticacion(navegador.driver)
arhivoDatos = NotepadCreator('usuarios.txt')
               
coordinador = CoordinadorAutenticacion(navegador, autenticacion, proxy_tester,arhivoDatos)

# Contraseña fija para el registro
contraseña = "Americano12."

# Proceso de autenticación usando proxies
coordinador.autenticar(contraseña)